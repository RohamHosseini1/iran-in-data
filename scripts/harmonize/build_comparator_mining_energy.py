#!/usr/bin/env python3
"""
Build comparator (+ Iran) mining / energy / industrial-output series from the USGS
Minerals Yearbook country "Table 1" workbooks already in data/raw/usgs-minerals-yearbook/.

Why USGS: it publishes the SAME table, with the same commodity definitions, for every
country in this project's comparator roster. That makes Iran and the comparators
methodologically comparable, which is the whole point of the comparator layer.

Stage 1 — extract: parse every country/edition "Table 1" sheet into tidy rows, using the
  cell indent level to reconstruct the commodity hierarchy (USGS nests sub-commodities
  under colon-terminated group headers; a naive parser mis-attributes them).
Stage 2 — dedup: a later edition's value for the same (commodity, year) supersedes an
  earlier one (USGS revises). Never averaged, never interpolated.
Stage 3 — harmonize: map the source's own (drifting) commodity labels onto a small set of
  canonical indicator_ids matching Iran's existing charts. The source's exact label is
  KEPT in `source_commodity_label` so nothing is silently reinterpreted.

Outputs (all under data/processed/comparator_mining_energy_series/):
  usgs_mineral_energy_production_long.csv   — harmonized, canonical indicator_ids
  usgs_mineral_energy_production_full.csv   — the complete deduped extraction (all commodities)

NO fabrication: values are copied through as published; USGS's own flags (e=estimated,
r=revised) are carried in the `flag` column. Blank/withheld cells are omitted, never zeroed.
"""
import csv
import os
import re
import sys

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RAW = os.path.join(REPO, "data", "raw", "usgs-minerals-yearbook")
OUT_DIR = os.path.join(REPO, "data", "processed", "comparator_mining_energy_series")

CATEGORY_LABELS = {"METALS", "INDUSTRIAL MINERALS", "MINERAL FUELS AND RELATED MATERIALS"}

# (country_iso3, dir, filename, edition_label). Edition order matters: later supersedes earlier.
SOURCE_FILES = [
    ("IRN", "irn", "myb3-2017-18-iran.xlsx", "MYB3-2017-18"),
    ("IRN", "irn", "myb3-2019-iran.xlsx", "MYB3-2019"),
    ("IRN", "irn", "myb3-2020-21-iran.xlsx", "MYB3-2020-21"),
    ("IRN", "irn", "myb3-2022-iran.xlsx", "MYB3-2022"),
    ("IRN", "irn", "myb3-2023-iran.xlsx", "MYB3-2023"),
    ("TUR", "tur", "myb3-2016-tu.xlsx", "MYB3-2016"),
    ("TUR", "tur", "myb3-2017-18-turkey.xlsx", "MYB3-2017-18"),
    ("TUR", "tur", "myb3-2019-turkey.xlsx", "MYB3-2019"),
    ("TUR", "tur", "myb3-2020-21-Turkey-advrel.xlsx", "MYB3-2020-21"),
    ("TUR", "tur", "myb3-2022-Turkey-ert.xlsx", "MYB3-2022"),
    ("SAU", "sau", "myb3-2017-18-saudi-arabia.xlsx", "MYB3-2017-18"),
    ("SAU", "sau", "myb3-2019-saudi-arabia.xlsx", "MYB3-2019"),
    ("SAU", "sau", "myb3-2020-21-saudi-arabia.xlsx", "MYB3-2020-21"),
    ("SAU", "sau", "myb3-2022-saudi-arabia.xlsx", "MYB3-2022"),
    ("SAU", "sau", "myb3-2023-saudi-arabia.xlsx", "MYB3-2023"),
    ("IRQ", "irq", "myb3-2017-18-iraq.xlsx", "MYB3-2017-18"),
    ("IRQ", "irq", "myb3-2019-iraq.xlsx", "MYB3-2019"),
    ("IRQ", "irq", "myb3-2020-21-iraq.xlsx", "MYB3-2020-21"),
    ("IRQ", "irq", "myb3-2022-Iraq-advrel.xlsx", "MYB3-2022"),
    ("IRQ", "irq", "myb3-2023-Iraq-ERT.xlsx", "MYB3-2023"),
    ("VEN", "ven", "myb3-2017-18-venezuela.xlsx", "MYB3-2017-18"),
    ("VEN", "ven", "myb3-2019-venezuela.xlsx", "MYB3-2019"),
    ("VEN", "ven", "myb3-2020-21-venezuela.xlsx", "MYB3-2020-21"),
    ("VEN", "ven", "myb3-2022-venezuela.xlsx", "MYB3-2022"),
    ("ARG", "arg", "myb3-2016-ar.xlsx", "MYB3-2016"),
    ("ARG", "arg", "myb3-2017-18-argentina.xlsx", "MYB3-2017-18"),
    ("ARG", "arg", "myb3-2019-argentina.xlsx", "MYB3-2019"),
    ("ARG", "arg", "myb3-2020-21-Argentina-ert.xlsx", "MYB3-2020-21"),
    ("RUS", "rus", "myb3-2017-18-russia.xlsx", "MYB3-2017-18"),
    ("RUS", "rus", "myb3-2019-russia.xlsx", "MYB3-2019"),
    ("RUS", "rus", "myb3-2020-21-russia.xlsx", "MYB3-2020-21"),
    ("RUS", "rus", "myb3-2022-russia.xlsx", "MYB3-2022"),
    ("KOR", "kor", "myb3-2017-18-korea-south.xlsx", "MYB3-2017-18"),
    ("KOR", "kor", "myb3-2019-south-korea.xlsx", "MYB3-2019"),
    ("KOR", "kor", "myb3-2020-21-republic-korea.xlsx", "MYB3-2020-21"),
    ("ESP", "esp", "myb3-2017-18-spain.xlsx", "MYB3-2017-18"),
    ("ESP", "esp", "myb3-2019-spain.xlsx", "MYB3-2019"),
    ("ESP", "esp", "myb3-2020-21-Spain-advrel.xlsx", "MYB3-2020-21"),
    ("ESP", "esp", "myb3-2022-Spain-ert.xlsx", "MYB3-2022"),
    ("ITA", "ita", "myb3-2017-18-italy.xlsx", "MYB3-2017-18"),
    ("ITA", "ita", "myb3-2019-italy.xlsx", "MYB3-2019"),
    ("ITA", "ita", "myb3-2022-Italy-ert.xlsx", "MYB3-2022"),
]

EDITION_ORDER = {
    "MYB3-2016": 0, "MYB3-2017-18": 1, "MYB3-2019": 2,
    "MYB3-2020-21": 3, "MYB3-2022": 4, "MYB3-2023": 5,
}

# ---------------------------------------------------------------- stage 1: extract


def strip_footnote_suffix(label):
    """USGS glues footnote markers onto label text: 'Natural gas, dry basis4'."""
    return re.sub(r'(?<=[a-zA-Z])\d+$', '', label).strip()


def parse_table1(path, country_iso3, edition_label):
    wb = openpyxl.load_workbook(path, data_only=True)
    # sheet name drifts across editions: 'Table 1', 'Table1', 'Table 1 '
    sheet = next((s for s in wb.sheetnames if s.strip().replace(" ", "").lower() == "table1"), None)
    if sheet is None:
        return []
    ws = wb[sheet]

    out = []
    category = ""
    # The table's DEFAULT unit (what a blank unit cell means) is declared in a parenthetical line
    # under the title, and it VARIES BY COUNTRY AND EDITION: Iran and Türkiye publish
    # "(Metric tons, gross weight, unless otherwise specified)" while Saudi Arabia and Iraq publish
    # "(Thousand metric tons, ...)". Hardcoding one default silently makes those countries 1000x
    # too small. It is read from the sheet, per table.
    default_unit = "metric tons, gross weight"
    last_unit = default_unit
    parent_stack = []  # [(indent_level, label)] of open colon-terminated group headers
    year_cols = []     # [(col_idx, year)]

    for cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row = [c.value for c in cells]
        c0 = cells[0]
        if c0.value is None:
            continue
        label = str(c0.value).strip()
        if not label:
            continue
        indent = c0.alignment.indent if (c0.alignment and c0.alignment.indent is not None) else 0

        # e.g. "(Thousand metric tons, gross weight, unless otherwise specified)"
        m = re.match(r'^\((.+?),?\s*unless otherwise (specified|noted)\.?\)?$', label, re.IGNORECASE)
        if m:
            default_unit = m.group(1).strip().rstrip(",").strip()
            last_unit = default_unit
            continue

        if label.startswith("Commodity"):
            # year headers are ints in some editions, strings in others
            year_cols = []
            for i, v in enumerate(row):
                if v is None:
                    continue
                if re.fullmatch(r'(19|20)\d{2}', str(v).strip()):
                    year_cols.append((i, int(str(v).strip())))
            parent_stack, last_unit = [], default_unit
            continue
        if label in CATEGORY_LABELS:
            category = label
            parent_stack, last_unit = [], default_unit
            continue
        if (label.startswith("TABLE")
                or label.upper().endswith("PRODUCTION OF MINERAL COMMODITIES1")
                or label.startswith("(")
                or label.lower().startswith("see footnotes")
                or re.match(r'^[a-zA-Z]?(Estimated|Revised|Ditto)\.?\s*$', label)
                or re.match(r'^\d[A-Za-z]', label)):   # footnote definition lines
            continue
        if not year_cols:
            continue

        while parent_stack and parent_stack[-1][0] >= indent:
            parent_stack.pop()

        has_data = any(row[i] is not None and str(row[i]).strip() != "" for i, _ in year_cols)
        base = label[:-1].strip() if label.endswith(":") else label
        base = strip_footnote_suffix(base)

        if label.endswith(":") and not has_data:
            # A group header ("Iron and steel:") carries no unit and must NOT disturb the
            # ditto reference -- the next row's 'do.' still points at the last DATA row's unit.
            parent_stack.append((indent, base))
            continue
        if not has_data:
            continue

        # Unit column semantics in USGS Table 1. Getting this wrong silently yields values off
        # by 1000x, or a tonnage series mislabelled 'kilograms' inherited from an unrelated gold
        # row above it -- so it is resolved explicitly, and only on data rows:
        #   explicit text -> that unit
        #   'do.'         -> ditto: the unit of the previous DATA row
        #   blank         -> the table's default unit, declared in the sheet header
        #                    ("Metric tons, gross weight, unless otherwise specified")
        unit_cell = row[2] if len(row) > 2 and row[2] is not None else None
        unit_text = str(unit_cell).strip() if unit_cell is not None else ""
        if unit_text == "do.":
            pass                       # keep last_unit
        elif unit_text:
            last_unit = unit_text
        else:
            last_unit = default_unit   # blank => table default, NOT the previous row's unit

        commodity = ", ".join([p[1] for p in parent_stack] + [base])

        for col, year in year_cols:
            v = row[col] if col < len(row) else None
            flag = row[col + 1] if col + 1 < len(row) else None
            if v is None or str(v).strip() == "":
                continue
            try:
                num = float(str(v).replace(",", ""))
            except (ValueError, TypeError):
                continue  # '--' (zero reported), 'NA', 'W' (withheld) -> not a number, skip
            out.append({
                "country_iso3": country_iso3,
                "category": category,
                "commodity": commodity,
                "unit": last_unit,
                "year": year,
                "value": int(num) if num == int(num) else num,
                "flag": str(flag).strip() if flag else "",
                "source_yearbook_edition": edition_label,
            })
    return out


# ------------------------------------------------- stage 3: canonical indicator mapping
#
# Each entry: indicator_id -> list of regexes matched (case-insensitively, fullmatch) against
# the source commodity label. Order matters: first indicator whose pattern matches wins.
# Deliberately CONSERVATIVE — a label we are not confident about is left unmapped rather than
# forced into an indicator it might not mean.

INDICATOR_PATTERNS = [
    # --- mining / minerals (align with Iran's existing per-commodity charts) ---
    ("mining__chromite_ore_production", [
        r"Chromium, mine, chromite, Ore",
        r"Chromium, mine, chromite(, concentrates?)?, (ores and concentrates, )?marketable",
        r"Chromium, (mine, chromite|Chromite, mine production), 34% to 43% Cr2O",
    ]),
    ("mining__copper_mine_production_cu_content", [
        r"Copper, Mine, (Ore, 0\.6% to 1\.2% Cu, )?Cu content",
        r"Copper, Mine, ore, 0\.6% to 1\.2% Cu, Cu content",
        r"Copper, Mine, Cu content, Ore",
        r"Copper, Mine, concentrates?, Cu content",
        r"Copper, Mine, Concentrates, Cu content",
        r"Copper, Mine, concentrates?, exclusive of pyrites?, Cu content",
        r"Copper, Mine production, exclusive of pyrite, Cu content",
        r"Copper, mine, concentrates, Cu content",
    ]),
    ("mining__copper_smelter_primary", [
        r"Copper, Smelter(, blister)?, Primary",
        r"Copper, Smelter production, Primary",
    ]),
    ("mining__copper_refinery_primary", [
        r"Copper, Refinery, Primary",
        r"Copper, Refinery, Primary, Total",
        r"Copper, Refinery production, Primary",
    ]),
    ("mining__iron_ore_production_gross_weight", [
        r"Iron ore, mine, Gross weight",
        r"Iron ore, mine production, Gross weight",
        r"Iron ore, mine, concentrates?, Gross weight",
        r"Iron ore, Gross weight",
    ]),
    ("mining__pig_iron_production", [
        r"Iron and steel, Pig iron",
    ]),
    ("mining__crude_steel_production", [
        r"Iron and steel, Raw steel",
        r"Iron and steel, Raw steel, including castings",
        r"Iron and steel, Raw steel, ingots,? and castings",
        r"Iron and steel, Steel, [Rr]aw( steel)?",
        r"Iron and steel, Steel, raw steel, ingots and castings",
        r"Steel, Raw steel",
    ]),
    ("mining__aluminum_primary_production", [
        r"Aluminum, Metals?, [Pp]rimary",
        r"Aluminum, metal, primary",
        r"Aluminum, primary",
        r"Aluminum, Aluminum, metal, primary",
    ]),
    ("mining__lead_mine_production_pb_content", [
        r"Lead, Mine, Pb content",
        r"Lead, Mine, concentrate, Pb content",
        r"Lead, Mine, Concentrate, Pb content",
        r"Lead, Mine production, Concentrate, Pb content",
        r"Lead, Mine, recoverable, Pb content",
        r"Lead, Pb content",
    ]),
    ("mining__zinc_mine_production_zn_content", [
        r"Zinc, Mine, Zn content",
        r"Zinc, mine, Zn content",
        r"Zinc, Mine, concentrate, Zn content",
        r"Zinc, mine, concentrate, Zn content",
        r"Zinc, mine production, Zn content",
        r"Zinc, Zn content",
    ]),
    ("mining__manganese_ore_production_gross_weight", [
        r"Manganese, mine, Gross weight",
        r"Manganese, mine production, Gross weight",
        r"Manganese, mine, concentrate(, marketable)?, Gross weight",
    ]),
    ("mining__barite_production", [
        r"Barite",
        r"Barite, crude and ground",
    ]),
    ("mining__gypsum_production", [
        r"Gypsum",
        r"Gypsum, mine",
        r"Gypsum, crude",
        r"Gypsum, including anhydrite",
    ]),
    ("mining__salt_production", [
        r"Salt",
        r"Salt, all types",
        r"Salt, common",
    ]),
    ("mining__cement_production", [
        r"Cement, [Hh]ydraulic",
    ]),
    ("mining__coke_production", [
        r"Coke, metallurgical",
        r"Coke, metallurgical, 6% moisture content",
    ]),
    # --- energy ---
    # Natural gas: USGS reports two GENUINELY DIFFERENT concepts and they must not be merged.
    #   gross      = total wellhead production, including gas reinjected/flared/vented
    #   marketed   = 'dry basis' / 'marketable' / unqualified 'Natural gas' — net of the above
    # Merging them would silently overstate any country reporting on a gross basis.
    ("energy__natural_gas_production_gross", [
        r"Natural gas, Gross",
    ]),
    ("energy__natural_gas_production_marketed", [
        r"Natural gas",
        r"Natural gas, dry basis",
        r"Natural gas, Dry basis",
        r"Natural gas, marketable",
        r"Natural gas, Marketable",
    ]),
    ("energy__crude_oil_production", [
        r"Petroleum, Crude",
        r"Petroleum, Crude, including condensate",
    ]),
    ("energy__refinery_output_total", [
        r"Petroleum, Refinery",
        r"Petroleum, [Rr]efinery, Total",
        r"Petroleum, Refinery production, Total",
        r"Petroleum, Refinery products",
    ]),
    ("energy__refinery_throughput", [
        r"Petroleum, Refinery, throughput",
    ]),
]

COMPILED = [(iid, [re.compile(p + r"e?", re.IGNORECASE) for p in pats])
            for iid, pats in INDICATOR_PATTERNS]
# trailing 'e?' absorbs USGS's glued 'e' (estimated) marker, e.g. 'Cement, hydraulice'


def map_indicator(commodity):
    for iid, pats in COMPILED:
        for p in pats:
            if p.fullmatch(commodity):
                return iid
    return None


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # stage 1
    raw = []
    for iso, d, fn, ed in SOURCE_FILES:
        path = os.path.join(RAW, d, fn)
        if not os.path.exists(path):
            print(f"MISSING {path}", file=sys.stderr)
            continue
        got = parse_table1(path, iso, ed)
        print(f"{iso} {ed:14s} {fn:42s} {len(got):5d} rows", file=sys.stderr)
        raw.extend(got)

    # stage 2 — dedup, later edition wins
    best = {}
    for r in raw:
        key = (r["country_iso3"], r["category"], r["commodity"], r["year"])
        if key not in best or EDITION_ORDER[r["source_yearbook_edition"]] >= EDITION_ORDER[best[key]["source_yearbook_edition"]]:
            best[key] = r
    full = sorted(best.values(), key=lambda r: (r["country_iso3"], r["category"], r["commodity"], r["year"]))

    full_path = os.path.join(OUT_DIR, "usgs_mineral_energy_production_full.csv")
    with open(full_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["country_iso3", "category", "commodity", "unit",
                                          "year", "value", "flag", "source_yearbook_edition"])
        w.writeheader()
        w.writerows(full)
    print(f"\nfull extraction: {len(raw)} raw -> {len(full)} deduped -> {full_path}", file=sys.stderr)

    # stage 3 — harmonize
    mapped = []
    for r in full:
        iid = map_indicator(r["commodity"])
        if not iid:
            continue
        mapped.append({
            "country_iso3": r["country_iso3"],
            "indicator_id": iid,
            "year": r["year"],
            "value": r["value"],
            "unit": r["unit"],
            "source_dataset": "usgs-minerals-yearbook/" + r["source_yearbook_edition"],
            "source_commodity_label": r["commodity"],
            "flag": r["flag"],
            "_ed": EDITION_ORDER[r["source_yearbook_edition"]],
        })

    # Several source labels can map to ONE indicator for the same country-year. Two causes:
    #   (a) label drift across editions for the same underlying series -- e.g. USGS renamed Iran's
    #       'Iron and steel, Raw steel, ingots and castings' to '... Steel, raw steel, ingots and
    #       castings'. Resolve by latest-edition-wins (same rule already used for value revisions).
    #   (b) two DIFFERENT labels present in the SAME edition -- that would be a real concept clash,
    #       not drift, and must not be silently collapsed. We detect it and report it.
    from collections import defaultdict
    groups = defaultdict(list)
    for r in mapped:
        groups[(r["country_iso3"], r["indicator_id"], r["year"])].append(r)

    conflicts = []
    long_rows = []
    for key, rs in groups.items():
        top = max(r["_ed"] for r in rs)
        winners = [r for r in rs if r["_ed"] == top]
        labels = {r["source_commodity_label"] for r in winners}
        if len(labels) > 1:
            conflicts.append((key, sorted(labels)))
            continue  # do NOT guess between genuinely competing labels; drop and report
        long_rows.append(winners[0])

    for r in long_rows:
        del r["_ed"]

    # ---- unit normalisation + dimensional quarantine -------------------------------------
    # Every indicator has ONE physical dimension. USGS's own sheets contain unit errors: e.g. in
    # myb3-2020-21-saudi-arabia.xlsx the 'Crude' petroleum row carries 'do.' (ditto) directly under
    # an 'Ethane / million cubic meters' row, which would make Saudi crude oil 3,635 million CUBIC
    # METRES. It is plainly a typesetting slip in the source (the value continues the country's own
    # million-42-gallon-barrel series), but it is not this project's place to silently rewrite a
    # published unit. So: rows whose resolved unit does not match the indicator's dimension are
    # QUARANTINED to unit_anomalies.csv -- kept, reported, and excluded from the chart-ready file.
    MASS_TO_T = {
        "metric tons": 1.0,
        "metric tons, gross weight": 1.0,
        "thousand metric tons": 1e3,
        "thousand metric tons, gross weight": 1e3,
        "thousand metrtic tons": 1e3,   # sic - typo present in the Iran 2023 sheet
        "million metric tons": 1e6,
        "million metric tons, gross weight": 1e6,
        "kilograms": 1e-3,
    }
    GAS_TO_MCM = {
        "million cubic meters": 1.0,
        "millions cubic meters": 1.0,   # sic
        "thousand cubic meters": 1e-3,
        "billion cubic meters": 1e3,
    }
    DIMENSION = {}
    for iid, _ in INDICATOR_PATTERNS:
        if iid.startswith("mining__"):
            DIMENSION[iid] = ("mass", "metric tons", MASS_TO_T)
        elif iid.startswith("energy__natural_gas"):
            DIMENSION[iid] = ("gas_volume", "million cubic meters", GAS_TO_MCM)
        else:
            DIMENSION[iid] = ("oil", None, None)   # see below

    normalised, anomalies = [], []
    for r in long_rows:
        dim, canon_unit, table = DIMENSION[r["indicator_id"]]
        if dim == "oil":
            # Crude-oil / refinery output are DELIBERATELY not published to the chart-ready file:
            #  1. the fuels section is exactly where the source-side 'do.' unit errors above occur,
            #     so the unit of a given cell cannot be verified from the sheet alone; and
            #  2. crude oil production is ALREADY covered for all 11 roster countries, from a single
            #     consistent source, by the existing chart owid__oil_production_volume.
            # The rows remain in usgs_mineral_energy_production_full.csv with their verbatim units.
            anomalies.append({**r, "anomaly": "oil/refinery series withheld: unverifiable unit in source fuels section"})
            continue
        factor = table.get(r["unit"].lower()) if r["unit"] else None
        if factor is None:
            anomalies.append({**r, "anomaly": f"unit '{r['unit']}' is not a {dim} unit for this indicator"})
            continue
        normalised.append({
            "country_iso3": r["country_iso3"],
            "indicator_id": r["indicator_id"],
            "year": r["year"],
            "value": round(r["value"] * factor, 4),
            "unit": canon_unit,
            "source_dataset": r["source_dataset"],
            "source_commodity_label": r["source_commodity_label"],
            "source_unit": r["unit"],
            "flag": r["flag"],
        })

    long_rows = sorted(normalised, key=lambda r: (r["indicator_id"], r["country_iso3"], r["year"]))
    anomalies.sort(key=lambda r: (r["indicator_id"], r["country_iso3"], r["year"]))

    anom_path = os.path.join(OUT_DIR, "unit_anomalies.csv")
    with open(anom_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["country_iso3", "indicator_id", "year", "value", "unit",
                                          "source_dataset", "source_commodity_label", "flag", "anomaly"])
        w.writeheader()
        w.writerows(anomalies)
    print(f"quarantined {len(anomalies)} rows -> {anom_path}", file=sys.stderr)

    if conflicts:
        print(f"\nWARNING: {len(conflicts)} country-year(s) had >1 competing source label within the "
              f"same edition; dropped rather than guessed:", file=sys.stderr)
        for key, labels in sorted(conflicts)[:20]:
            print(f"  {key}: {labels}", file=sys.stderr)

    long_path = os.path.join(OUT_DIR, "usgs_mineral_energy_production_long.csv")
    with open(long_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["country_iso3", "indicator_id", "year", "value", "unit",
                                          "source_dataset", "source_commodity_label", "source_unit", "flag"])
        w.writeheader()
        w.writerows(long_rows)
    print(f"harmonized: {len(long_rows)} rows -> {long_path}", file=sys.stderr)

    # coverage report to stdout (used to write the README honestly)
    from collections import defaultdict
    cov = defaultdict(lambda: defaultdict(list))
    for r in long_rows:
        cov[r["indicator_id"]][r["country_iso3"]].append(r["year"])
    print("\nindicator_id,country_iso3,n_years,min_year,max_year")
    for iid in sorted(cov):
        for iso in sorted(cov[iid]):
            ys = cov[iid][iso]
            print(f"{iid},{iso},{len(ys)},{min(ys)},{max(ys)}")


if __name__ == "__main__":
    main()
